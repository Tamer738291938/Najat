# -*- coding: utf-8 -*-
"""
تحديث بيانات لوحة متابعة الخطة التشغيلية من ملف الإكسل.

يقرأ: Docs/سجل_متابعة_الخطة_التشغيلية_2026-2027.xlsx
يحدّث: Dashboard/najat_dashboard.html (مصفوفات ACTS / KPIS / DEPS / RISKS + تاريخ آخر تحديث)

الاستخدام:  python update_dashboard.py
المتطلبات:  pip install openpyxl
"""
import json, re, sys, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
HTML = ROOT / "Dashboard" / "najat_dashboard.html"

def find_xlsx():
    hits = sorted((ROOT / "Docs").glob("سجل_متابعة*.xlsx"),
                  key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        sys.exit("لم يُعثر على ملف سجل المتابعة (سجل_متابعة*.xlsx) داخل مجلد Docs")
    return hits[0]  # الأحدث تعديلاً

AR_MONTHS = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
             "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

def cell(v):
    if v is None: return ""
    return str(v).strip()

def num(v):
    """قيمة رقمية أو None (الخلايا الفارغة والشرطات تعني: لا قيمة)."""
    if v is None or cell(v) in ("", "—", "-"): return None
    try: return float(v)
    except (TypeError, ValueError): return None

def js(rows):
    out = ",\n".join(json.dumps(r, ensure_ascii=False) for r in rows)
    return "[\n" + out + "\n]"

def main():
    xlsx = find_xlsx()
    print(f"ملف المصدر: {xlsx.name}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    acts = []
    for r in wb["الأنشطة"].iter_rows(min_row=2, values_only=True):
        if num(r[0]) is None: continue
        acts.append([int(r[0]), cell(r[1]), cell(r[2]), cell(r[3]),
                     cell(r[4]), cell(r[5]), cell(r[6]) or "لم يبدأ"])

    kpis = []
    for r in wb["المؤشرات"].iter_rows(min_row=2, values_only=True):
        if not cell(r[0]).startswith("KPI"): continue
        note = cell(r[12]) if len(r) > 12 else ""
        is_count = cell(r[1]).startswith("عدد")
        if is_count and "عدد" not in note:
            note = (note + " · " if note else "") + "عدد (وليس نسبة)"
        kpis.append([cell(r[0]), cell(r[1]), cell(r[2]), cell(r[3]), cell(r[4]),
                     num(r[6]), num(r[9]),          # مستهدف فصل1، مستهدف نهاية العام
                     num(r[7]), num(r[10]),         # فعلي فصل1، فعلي نهاية العام
                     cell(r[8]) or "—", cell(r[11]) or "—", note])

    deps = []
    for r in wb["التبعيات الحرجة"].iter_rows(min_row=2, values_only=True):
        if num(r[0]) is None: continue
        deps.append([int(r[0]), cell(r[1]), cell(r[2]), cell(r[3]),
                     cell(r[4]) or "لم يبدأ", cell(r[5])])

    risks = []
    for r in wb["المخاطر"].iter_rows(min_row=2, values_only=True):
        if not cell(r[0]).startswith("R"): continue
        risks.append([cell(r[0]), cell(r[1]), cell(r[2]), cell(r[3]),
                      cell(r[4]), cell(r[5]) or "مفتوح"])

    html = HTML.read_text(encoding="utf-8")
    for name, rows in [("ACTS", acts), ("KPIS", kpis), ("DEPS", deps), ("RISKS", risks)]:
        pat = re.compile(r"const %s=\[.*?\n\];" % name, re.S)
        if not pat.search(html):
            sys.exit(f"لم يُعثر على مصفوفة {name} في ملف اللوحة")
        html = pat.sub(lambda m: f"const {name}={js(rows)};", html, count=1)

    today = datetime.date.today()
    stamp = f"{AR_MONTHS[today.month-1]} {today.year}"
    html = re.sub(r'آخر تحديث للبيانات <b>[^<]*</b>',
                  f'آخر تحديث للبيانات <b>{stamp}</b>', html)

    HTML.write_text(html, encoding="utf-8")
    print(f"تم التحديث: {len(acts)} نشاطاً، {len(kpis)} مؤشراً، "
          f"{len(deps)} تبعيات، {len(risks)} مخاطر — بتاريخ {stamp}")

if __name__ == "__main__":
    main()
